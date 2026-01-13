
import os
import sys
import re
import math
import threading
from dataclasses import dataclass
from typing import Dict, Optional, Tuple, List

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pandas as pd
import numpy as np
import openpyxl

from reportlab.lib.pagesizes import LETTER
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet


# ============================================================
# CONFIG
# ============================================================
SCHEMA_FILENAME_DEFAULT = "Comisiones esquema 2026 2.0 (1).xlsm"

BASE_SHEET_VENTAS = "ResultadosVentascomisionesporc"
BASE_SHEET_FILTRO = "Hoja2"

IVA_FACTOR = 1.16

# Excluir líneas de impuestos/IVA por nombre de producto
NON_COMMISSION_PRODUCT_REGEX = re.compile(r"(?i)\b(iva|i\.?v\.?a\.?|impuesto|tax)\b")

# UI sizing
DEFAULT_GEOMETRY = "900x600"
MIN_W, MIN_H = 860, 560

# Paginación
DEFAULT_PAGE_SIZE = 200
PAGE_SIZES = [50, 100, 200, 500, 1000]

# Columnas de detalle en la app (SIN TIPO)
DISPLAY_COLS = [
    "Fecha",
    "Asesor",
    "Cliente",
    "OV",
    "Producto",
    "Cantidad",
    "Precio Bruto",
    "Precio Unitario Neto",
    "Venta Total",
    "Precio 4",
    "Precio 3",
    "Precio 2",
    "Precio 1",
    "Comisión",
    "Total comisión",
]


# ============================================================
# Helpers
# ============================================================
def app_base_dir() -> str:
    # Para PyInstaller
    return getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))

def default_schema_path() -> str:
    return os.path.join(app_base_dir(), SCHEMA_FILENAME_DEFAULT)

def center_window(win: tk.Tk):
    win.update_idletasks()
    w = win.winfo_width()
    h = win.winfo_height()
    sw = win.winfo_screenwidth()
    sh = win.winfo_screenheight()
    x = max(0, (sw - w) // 2)
    y = max(0, (sh - h) // 2)
    win.geometry(f"{w}x{h}+{x}+{y}")

def safe_float(x, default=float("nan")):
    try:
        if pd.isna(x):
            return default
        return float(x)
    except Exception:
        return default

def money(x):
    try:
        if pd.isna(x):
            return ""
        return f"{float(x):,.2f}"
    except Exception:
        return str(x)

def pct(x):
    try:
        if pd.isna(x):
            return ""
        return f"{float(x) * 100:.4f}%"
    except Exception:
        return str(x)

def norm_ov(v) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return str(v).strip()
    return str(v).strip()

def is_tax_line(product_value) -> bool:
    s = "" if product_value is None else str(product_value).strip()
    if not s:
        return False
    return NON_COMMISSION_PRODUCT_REGEX.search(s) is not None

def norm_product_key(v) -> str:
    if v is None:
        return ""
    return str(v).strip().upper()


# ============================================================
# Ventas válidas (Hoja2: ov + cruce)
# ============================================================
def extract_valid_ovs_from_hoja2(xlsx_path: str) -> set[str]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if BASE_SHEET_FILTRO not in wb.sheetnames:
        raise ValueError(f"No encontré la hoja '{BASE_SHEET_FILTRO}' en la base_comisiones.")

    ws = wb[BASE_SHEET_FILTRO]

    header_row = None
    ov_col = None
    cruce_col = None
    for r in range(1, 80):
        vals = []
        for c in range(1, 60):
            v = ws.cell(r, c).value
            vals.append(v.strip().lower() if isinstance(v, str) else v)
        if "ov" in vals and "cruce" in vals:
            header_row = r
            ov_col = vals.index("ov") + 1
            cruce_col = vals.index("cruce") + 1
            break

    if header_row is None:
        raise ValueError("No pude ubicar encabezados 'ov' y 'cruce' en Hoja2.")

    valid = set()
    for r in range(header_row + 1, ws.max_row + 1):
        ov = norm_ov(ws.cell(r, ov_col).value)
        cr = norm_ov(ws.cell(r, cruce_col).value)
        # Regla: solo válidos si ambas columnas tienen valor
        if ov and cr:
            valid.add(ov)
            valid.add(cr)

    return valid


# ============================================================
# Reglas desde esquema fijo (COMISIONES 2026 + NUEVAS LISTAS)
# ============================================================
@dataclass(frozen=True)
class CommissionBracket:
    lim_inf: float
    lim_sup: float
    p4: float
    p3: float
    p2: float
    p1: float

class Rules2026:
    def __init__(self, schema_file: str):
        self.schema_file = schema_file
        self.comm_brackets: List[CommissionBracket] = []
        self.price_map: Dict[str, Dict[str, float]] = {}  # PRODUCT_KEY -> {p4,p3,p2,p1}
        self.tipo_map: Dict[str, str] = {}  # ASESOR_UPPER -> TIPO

    def load(self):
        if not os.path.exists(self.schema_file):
            raise FileNotFoundError(
                f"No encontré el esquema fijo:\n{self.schema_file}\n\n"
                f"Coloca el archivo '{os.path.basename(self.schema_file)}' junto al programa "
                "o selecciónalo con el botón de Esquema."
            )

        wb = openpyxl.load_workbook(self.schema_file, data_only=True, keep_vba=True)

        self._load_commissions(wb)
        self._load_prices(wb)
        self._load_tipos(wb)

    def _load_commissions(self, wb):
        if "COMISIONES 2026" not in wb.sheetnames:
            raise ValueError("No encontré la hoja 'COMISIONES 2026' en el esquema.")
        ws = wb["COMISIONES 2026"]

        header_row = None
        for r in range(1, 80):
            b = ws.cell(r, 2).value
            c = ws.cell(r, 3).value
            if isinstance(b, str) and isinstance(c, str):
                if b.strip().lower() == "limite inf" and c.strip().lower() == "limite sup":
                    header_row = r
                    break
        if header_row is None:
            raise ValueError("No pude localizar encabezados 'Limite inf'/'Limite sup' en COMISIONES 2026.")

        brackets: List[CommissionBracket] = []
        for r in range(header_row + 1, header_row + 600):
            li = ws.cell(r, 2).value
            ls = ws.cell(r, 3).value
            if li is None or ls is None:
                if brackets:
                    break
                continue

            p4 = ws.cell(r, 4).value
            p3 = ws.cell(r, 5).value
            p2 = ws.cell(r, 6).value
            p1 = ws.cell(r, 7).value

            try:
                brackets.append(CommissionBracket(
                    lim_inf=float(li),
                    lim_sup=float(ls),
                    p4=float(p4),
                    p3=float(p3),
                    p2=float(p2),
                    p1=float(p1),
                ))
            except Exception:
                continue

        if not brackets:
            raise ValueError("No pude leer la tabla de 'COMISIONES 2026'.")

        self.comm_brackets = sorted(brackets, key=lambda b: b.lim_inf)

    def _load_prices(self, wb):
        if "NUEVAS LISTAS" not in wb.sheetnames:
            raise ValueError("No encontré la hoja 'NUEVAS LISTAS' en el esquema.")
        ws = wb["NUEVAS LISTAS"]

        header_row = None
        for r in range(1, 80):
            v = ws.cell(r, 2).value
            if isinstance(v, str) and v.strip().upper() == "MODELO":
                header_row = r
                break
        if header_row is None:
            raise ValueError("No pude localizar encabezado 'MODELO' en NUEVAS LISTAS.")

        # Según plantilla: +200(p4)=5, +100(p3)=8, +50(p2)=11, +20(p1)=14
        p4_col, p3_col, p2_col, p1_col = 5, 8, 11, 14

        price_map: Dict[str, Dict[str, float]] = {}
        for r in range(header_row + 1, ws.max_row + 1):
            modelo = ws.cell(r, 2).value
            modelo_s = (str(modelo).strip() if modelo is not None else "")
            if not modelo_s:
                continue
            key = modelo_s.upper()

            p4 = ws.cell(r, p4_col).value
            p3 = ws.cell(r, p3_col).value
            p2 = ws.cell(r, p2_col).value
            p1 = ws.cell(r, p1_col).value

            if all(v is None for v in (p4, p3, p2, p1)):
                continue

            price_map[key] = {
                "p4": safe_float(p4),
                "p3": safe_float(p3),
                "p2": safe_float(p2),
                "p1": safe_float(p1),
            }

        if not price_map:
            raise ValueError("No pude construir la lista de productos desde NUEVAS LISTAS.")

        self.price_map = price_map


    def _load_tipos(self, wb):
        # Opcional: mapeo de asesor -> tipo desde hoja "Catalogos"
        if "Catalogos" not in wb.sheetnames:
            self.tipo_map = {}
            return
        ws = wb["Catalogos"]

        header_row = None
        col_asesor = None
        col_tipo = None
        for r in range(1, 80):
            b = ws.cell(r, 2).value
            c = ws.cell(r, 3).value
            if isinstance(b, str) and isinstance(c, str):
                if b.strip().upper() == "ASESORES" and c.strip().upper() == "TIPO":
                    header_row = r
                    col_asesor = 2
                    col_tipo = 3
                    break

        if header_row is None:
            self.tipo_map = {}
            return

        m = {}
        for r in range(header_row + 1, ws.max_row + 1):
            a = ws.cell(r, col_asesor).value
            t = ws.cell(r, col_tipo).value
            if a is None:
                continue
            a_key = str(a).strip().upper()
            if not a_key:
                continue
            m[a_key] = "" if t is None else str(t).strip()
        self.tipo_map = m


def pick_bracket(brackets: List[CommissionBracket], total_sales: float) -> CommissionBracket:
    """
    - Si total < mínimo limite inf: 0% en todos
    - Si total > máximo limite sup: usar último rango (clamp)
    """
    if not brackets:
        return CommissionBracket(0, float("inf"), 0, 0, 0, 0)

    if total_sales < brackets[0].lim_inf:
        return CommissionBracket(0, brackets[0].lim_inf, 0, 0, 0, 0)

    for b in brackets:
        if b.lim_inf <= total_sales <= b.lim_sup:
            return b

    return brackets[-1]


def infer_level(price_to_compare: float, p4: float, p3: float, p2: float, p1: float) -> int:
    """
    Regla pedida (clamp):
      - Si <= P4 -> nivel 4 (menor comisión)
      - Si <= P3 -> nivel 3
      - Si <= P2 -> nivel 2
      - Si <= P1 -> nivel 1
      - Si > P1  -> nivel 1
    Si falta info -> nivel 4 (conservador)
    """
    if any(pd.isna(v) for v in (price_to_compare, p4, p3, p2, p1)):
        return 4
    if price_to_compare <= p4:
        return 4
    if price_to_compare <= p3:
        return 3
    if price_to_compare <= p2:
        return 2
    return 1


def rate_for_level(bracket: CommissionBracket, level: int) -> float:
    if level == 1:
        return bracket.p1
    if level == 2:
        return bracket.p2
    if level == 3:
        return bracket.p3
    return bracket.p4


# ============================================================
# PDF Carátula
# (simple y consistente: usa el mismo dataset filtrado)
# ============================================================
def export_caratula_pdf(path_pdf: str,
                        resumen_df: pd.DataFrame,
                        fecha_ini: pd.Timestamp,
                        fecha_fin: pd.Timestamp,
                        fecha_pago: pd.Timestamp,
                        incluir_tipo: bool = False):
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(path_pdf, pagesize=LETTER, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)

    story = []
    titulo = f"CALCULO DE COMISIONES DEL {fecha_ini:%d-%b-%Y} AL {fecha_fin:%d-%b-%Y}"
    story.append(Paragraph(titulo.upper(), styles["Title"]))
    story.append(Spacer(1, 10))
    story.append(Paragraph(f"COMISIONES A PAGO {fecha_pago:%d-%b-%y}".upper(), styles["Normal"]))
    story.append(Spacer(1, 14))

    if incluir_tipo and "Tipo" in resumen_df.columns:
        header = ["NOMBRE ASESOR", "TIPO", "VENTAS", "TOTAL COMISION $"]
    else:
        header = ["NOMBRE ASESOR", "VENTAS", "TOTAL COMISION $"]

    data = [header]

    for _, r in resumen_df.iterrows():
        if incluir_tipo and "Tipo" in resumen_df.columns:
            data.append([
                str(r["Asesor"]),
                str(r.get("Tipo", "")),
                money(r["Venta Total"]),
                money(r["Total comisión"]),
            ])
        else:
            data.append([
                str(r["Asesor"]),
                money(r["Venta Total"]),
                money(r["Total comisión"]),
            ])

    # Totales
    if incluir_tipo and "Tipo" in resumen_df.columns:
        data.append(["TOTALES", "", money(resumen_df["Venta Total"].sum()), money(resumen_df["Total comisión"].sum())])
        col_widths = [220, 80, 120, 120]
    else:
        data.append(["TOTALES", money(resumen_df["Venta Total"].sum()), money(resumen_df["Total comisión"].sum())])
        col_widths = [280, 120, 120]

    table = Table(data, hAlign="LEFT", colWidths=col_widths)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.whitesmoke),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(table)
    doc.build(story)


# ============================================================
# App (Tkinter)
# ============================================================
class Comisionador2026App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Comisionador 2026 (Tkinter)")
        self.geometry(DEFAULT_GEOMETRY)
        self.minsize(MIN_W, MIN_H)

        # Estado
        self.rules: Optional[Rules2026] = None
        self.schema_file: str = default_schema_path()

        self.base_file: Optional[str] = None
        self.base_df_all: Optional[pd.DataFrame] = None  # dataset filtrado por OV/IVA/producto (sin fecha)
        self.audit_counts: Dict[str, int] = {}

        self.out_df: Optional[pd.DataFrame] = None
        self.resumen_df: Optional[pd.DataFrame] = None

        # Paginación
        self.page_size = DEFAULT_PAGE_SIZE
        self.current_page = 1
        self.total_pages = 1

        # UI layout: centrado y responsivo
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self._build_ui()
        self.after(50, lambda: center_window(self))

        # Cargar reglas al iniciar
        self._load_rules_initial()

    # ------------------ UI ------------------
    def _build_ui(self):
        root = ttk.Frame(self, padding=10)
        root.grid(row=0, column=0, sticky="nsew")
        root.grid_rowconfigure(1, weight=1)
        root.grid_columnconfigure(0, weight=1)

        # Top: controles
        top = ttk.LabelFrame(root, text="Flujo", padding=10)
        top.grid(row=0, column=0, sticky="ew")
        top.grid_columnconfigure(0, weight=1)

        # Row 0: botones (no ocupan todo el ancho)
        bar = ttk.Frame(top)
        bar.grid(row=0, column=0, sticky="w")

        self.btn_schema = ttk.Button(bar, text="Seleccionar esquema (opcional)", width=28, command=self.pick_schema)
        self.btn_base = ttk.Button(bar, text="Cargar base_comisiones (.xlsx)", width=28, command=self.pick_base)
        self.btn_process = ttk.Button(bar, text="Procesar", width=12, command=self.process_async, state="disabled")
        self.btn_export = ttk.Button(bar, text="Exportar Carátula PDF", width=22, command=self.export_pdf, state="disabled")

        self.btn_schema.grid(row=0, column=0, padx=(0, 8), pady=4)
        self.btn_base.grid(row=0, column=1, padx=(0, 8), pady=4)
        self.btn_process.grid(row=0, column=2, padx=(0, 8), pady=4)
        self.btn_export.grid(row=0, column=3, padx=(0, 8), pady=4)

        # Row 1: fechas + switches
        opts = ttk.Frame(top)
        opts.grid(row=1, column=0, sticky="w", pady=(6, 0))

        ttk.Label(opts, text="Inicio:").grid(row=0, column=0, padx=(0, 6), sticky="e")
        ttk.Label(opts, text="Fin:").grid(row=0, column=2, padx=(12, 6), sticky="e")
        ttk.Label(opts, text="Pago:").grid(row=0, column=4, padx=(12, 6), sticky="e")

        self.cb_ini = ttk.Combobox(opts, state="disabled", width=12)
        self.cb_fin = ttk.Combobox(opts, state="disabled", width=12)
        self.cb_pago = ttk.Combobox(opts, state="disabled", width=12)

        self.cb_ini.grid(row=0, column=1, padx=(0, 6))
        self.cb_fin.grid(row=0, column=3, padx=(0, 6))
        self.cb_pago.grid(row=0, column=5, padx=(0, 6))

        self.var_filter_date = tk.BooleanVar(value=False)
        self.chk_filter_date = ttk.Checkbutton(
            opts,
            text="Filtrar ventas por Fecha (si no, solo etiqueta)",
            variable=self.var_filter_date
        )
        self.chk_filter_date.grid(row=1, column=0, columnspan=6, sticky="w", pady=(6, 0))

        self.var_compare_net = tk.BooleanVar(value=True)
        self.chk_compare_net = ttk.Checkbutton(
            opts,
            text="Comparar precios por Neto (PU*1.16) (si no, por Bruto/PU)",
            variable=self.var_compare_net
        )
        self.chk_compare_net.grid(row=2, column=0, columnspan=6, sticky="w", pady=(2, 0))

        self.var_pdf_tipo = tk.BooleanVar(value=False)
        self.chk_pdf_tipo = ttk.Checkbutton(
            opts,
            text="Incluir TIPO en carátula (si existe en esquema)",
            variable=self.var_pdf_tipo
        )
        self.chk_pdf_tipo.grid(row=3, column=0, columnspan=6, sticky="w", pady=(2, 0))

        # Tabs
        self.nb = ttk.Notebook(root)
        self.nb.grid(row=1, column=0, sticky="nsew", pady=(8, 0))

        self.tab_detalle = ttk.Frame(self.nb, padding=8)
        self.tab_resumen = ttk.Frame(self.nb, padding=8)
        self.tab_audit = ttk.Frame(self.nb, padding=8)

        self.nb.add(self.tab_detalle, text="Detalle")
        self.nb.add(self.tab_resumen, text="Carátula (Resumen)")
        self.nb.add(self.tab_audit, text="Auditoría")

        self._build_tables()

        # Bottom: status + progress
        bottom = ttk.Frame(root)
        bottom.grid(row=2, column=0, sticky="ew", pady=(8, 0))
        bottom.grid_columnconfigure(0, weight=1)

        self.status = tk.StringVar(value="Listo.")
        self.lbl_status = ttk.Label(bottom, textvariable=self.status)
        self.lbl_status.grid(row=0, column=0, sticky="w")

        self.pb = ttk.Progressbar(bottom, mode="indeterminate")
        self.pb.grid(row=1, column=0, sticky="ew", pady=(6, 0))

    def _build_tables(self):
        # ---- Detalle (paginación + tree) ----
        self.tab_detalle.grid_rowconfigure(1, weight=1)
        self.tab_detalle.grid_columnconfigure(0, weight=1)

        pager = ttk.Frame(self.tab_detalle)
        pager.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        pager.grid_columnconfigure(20, weight=1)

        ttk.Label(pager, text="Tamaño página:").grid(row=0, column=0, padx=(0, 6), sticky="w")
        self.cb_pagesize = ttk.Combobox(pager, state="readonly", width=6, values=[str(x) for x in PAGE_SIZES])
        self.cb_pagesize.set(str(DEFAULT_PAGE_SIZE))
        self.cb_pagesize.grid(row=0, column=1, padx=(0, 12), sticky="w")
        self.cb_pagesize.bind("<<ComboboxSelected>>", self.on_pagesize_change)

        self.btn_first = ttk.Button(pager, text="<<", width=4, command=self.go_first, state="disabled")
        self.btn_prev = ttk.Button(pager, text="<", width=4, command=self.go_prev, state="disabled")
        self.lbl_page = ttk.Label(pager, text="Página 0 / 0")
        self.btn_next = ttk.Button(pager, text=">", width=4, command=self.go_next, state="disabled")
        self.btn_last = ttk.Button(pager, text=">>", width=4, command=self.go_last, state="disabled")

        self.btn_first.grid(row=0, column=2, padx=2)
        self.btn_prev.grid(row=0, column=3, padx=2)
        self.lbl_page.grid(row=0, column=4, padx=10)
        self.btn_next.grid(row=0, column=5, padx=2)
        self.btn_last.grid(row=0, column=6, padx=2)

        frame = ttk.Frame(self.tab_detalle)
        frame.grid(row=1, column=0, sticky="nsew")
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        self.tree = ttk.Treeview(frame, columns=DISPLAY_COLS, show="headings")
        vs = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        hs = ttk.Scrollbar(frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vs.set, xscrollcommand=hs.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vs.grid(row=0, column=1, sticky="ns")
        hs.grid(row=1, column=0, sticky="ew")

        for c in DISPLAY_COLS:
            self.tree.heading(c, text=c)
            w = 115
            if c in ("Cliente", "Asesor"):
                w = 220
            if c in ("OV", "Producto"):
                w = 150
            if c in ("Cantidad",):
                w = 90
            self.tree.column(c, width=w, anchor="center", stretch=False)

        # ---- Resumen ----
        self.tab_resumen.grid_rowconfigure(0, weight=1)
        self.tab_resumen.grid_columnconfigure(0, weight=1)

        rframe = ttk.Frame(self.tab_resumen)
        rframe.grid(row=0, column=0, sticky="nsew")
        rframe.grid_rowconfigure(0, weight=1)
        rframe.grid_columnconfigure(0, weight=1)

        self.sum_cols = ["Asesor", "Venta Total", "Total comisión"]
        self.tree_sum = ttk.Treeview(rframe, columns=self.sum_cols, show="headings")
        v2 = ttk.Scrollbar(rframe, orient="vertical", command=self.tree_sum.yview)
        self.tree_sum.configure(yscrollcommand=v2.set)

        self.tree_sum.grid(row=0, column=0, sticky="nsew")
        v2.grid(row=0, column=1, sticky="ns")

        for c in self.sum_cols:
            self.tree_sum.heading(c, text=c)
            w = 260 if c == "Asesor" else 160
            self.tree_sum.column(c, width=w, anchor="center", stretch=False)

        # ---- Auditoría ----
        self.tab_audit.grid_rowconfigure(0, weight=1)
        self.tab_audit.grid_columnconfigure(0, weight=1)

        self.txt_audit = tk.Text(self.tab_audit, wrap="word")
        scr = ttk.Scrollbar(self.tab_audit, orient="vertical", command=self.txt_audit.yview)
        self.txt_audit.configure(yscrollcommand=scr.set)

        self.txt_audit.grid(row=0, column=0, sticky="nsew")
        scr.grid(row=0, column=1, sticky="ns")

    # ------------------ Progress helpers ------------------
    def _start_busy(self, msg: str):
        self.status.set(msg)
        self.pb.start(12)

    def _stop_busy(self, msg: str):
        self.pb.stop()
        self.status.set(msg)

    def _set_status(self, msg: str):
        self.status.set(msg)

    # ------------------ Rules ------------------
    def _load_rules_initial(self):
        try:
            self._start_busy("Cargando esquema fijo...")
            self.rules = Rules2026(self.schema_file)
            self.rules.load()
            self._stop_busy(f"Reglas cargadas: {os.path.basename(self.schema_file)}")
        except Exception as ex:
            self.rules = None
            self.pb.stop()
            self.status.set("No se pudieron cargar reglas. Selecciona el esquema.")
            messagebox.showwarning("Esquema", str(ex))

    def pick_schema(self):
        p = filedialog.askopenfilename(
            title="Selecciona el esquema (xlsm/xlsx)",
            filetypes=[("Excel", "*.xlsm *.xlsx"), ("Todos", "*.*")]
        )
        if not p:
            return
        self.schema_file = p
        try:
            self._start_busy("Cargando esquema...")
            self.rules = Rules2026(self.schema_file)
            self.rules.load()
            self._stop_busy(f"Reglas cargadas: {os.path.basename(self.schema_file)}")
            # si ya hay base cargada, habilita procesar
            if self.base_df_all is not None:
                self.btn_process.configure(state="normal")
        except Exception as ex:
            self.rules = None
            self._stop_busy("Error cargando esquema.")
            messagebox.showerror("Esquema", str(ex))

    # ------------------ Base ------------------
    def pick_base(self):
        if self.rules is None:
            messagebox.showerror("Reglas", "Primero carga el esquema (COMISIONES 2026 y NUEVAS LISTAS).")
            return

        p = filedialog.askopenfilename(
            title="Selecciona base_comisiones (.xlsx)",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")]
        )
        if not p:
            return

        self.base_file = p
        self.btn_base.configure(state="disabled")
        self.btn_process.configure(state="disabled")
        self.btn_export.configure(state="disabled")
        self._start_busy("Cargando base...")

        threading.Thread(target=self._load_base_thread, args=(p,), daemon=True).start()

    def _load_base_thread(self, path_xlsx: str):
        try:
            self._set_status("Leyendo Hoja2 (ov+cruce)...")
            valid_ovs = extract_valid_ovs_from_hoja2(path_xlsx)

            if not valid_ovs:
                raise ValueError("No se encontraron OVs válidas en Hoja2 (ov y cruce con valor).")

            self._set_status("Leyendo ResultadosVentascomisionesporc...")
            # Columnas por letra: A Fecha, D Rep, E Nombre, H Artículo, I Cantidad, S PU, T OV
            df = pd.read_excel(
                path_xlsx,
                sheet_name=BASE_SHEET_VENTAS,
                engine="openpyxl",
                usecols="A,D,E,H,I,S,T"
            )
            df = df.rename(columns={
                df.columns[0]: "Fecha",
                df.columns[1]: "Asesor",
                df.columns[2]: "Cliente",
                df.columns[3]: "Producto",
                df.columns[4]: "Cantidad",
                df.columns[5]: "Precio Bruto",  # PU (subtotal)
                df.columns[6]: "OV",
            })

            self._set_status("Aplicando filtros (OV válida / IVA / productos)...")
            df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
            df = df.dropna(subset=["Fecha"]).copy()

            # eliminar IVA/impuestos
            df = df[~df["Producto"].apply(is_tax_line)].copy()

            # Filtrar OVs válidas
            df["OV"] = df["OV"].apply(norm_ov)
            df = df[df["OV"].isin(valid_ovs)].copy()

            # Numéricos y nulos
            df["Cantidad"] = pd.to_numeric(df["Cantidad"], errors="coerce")
            df["Precio Bruto"] = pd.to_numeric(df["Precio Bruto"], errors="coerce")
            df = df.dropna(subset=["Producto", "Cantidad", "Precio Bruto"]).copy()
            df = df[df["Cantidad"] != 0].copy()

            # Producto en NUEVAS LISTAS
            df["Producto_key"] = df["Producto"].apply(norm_product_key)
            valid_products = set(self.rules.price_map.keys())
            df = df[df["Producto_key"].isin(valid_products)].copy()

            if df.empty:
                raise ValueError(
                    "Después de filtrar por OVs válidas + excluir IVA + productos en NUEVAS LISTAS, no quedaron filas.\n"
                    "Revisa que Producto en la base coincida con MODELO en NUEVAS LISTAS."
                )

            # Guardar dataset base (sin filtrar por fecha; fecha es etiqueta o filtro opcional)
            # Pre-cargar lista de fechas para dropdowns
            dates = sorted(df["Fecha"].dt.date.unique())
            vals = [d.strftime("%Y-%m-%d") for d in dates]

            # Auditoría
            audit = {
                "rows_valid_after_all_filters": int(len(df)),
                "unique_asesores": int(df["Asesor"].nunique()),
                "unique_ovs": int(df["OV"].nunique()),
                "date_min": str(min(dates)) if dates else "",
                "date_max": str(max(dates)) if dates else "",
            }

            def ui_apply():
                self.base_df_all = df
                self.audit_counts = audit

                for cb in (self.cb_ini, self.cb_fin, self.cb_pago):
                    cb.configure(state="readonly", values=vals)
                self.cb_ini.set(vals[0])
                self.cb_fin.set(vals[-1])
                self.cb_pago.set(vals[-1])

                # reset outputs
                self.out_df = None
                self.resumen_df = None
                self._clear_tables()
                self._reset_pagination()
                self._render_audit_text()

                self.btn_process.configure(state="normal")
                self.btn_base.configure(state="normal")
                self._stop_busy(f"Base lista. Filas válidas: {len(df):,}")

            self.after(0, ui_apply)

        except Exception as ex:
            def ui_err():
                self.btn_base.configure(state="normal")
                self.btn_process.configure(state="disabled")
                self.btn_export.configure(state="disabled")
                self._stop_busy("Error al cargar base.")
                messagebox.showerror("Base", str(ex))
            self.after(0, ui_err)

    # ------------------ Procesar ------------------
    def process_async(self):
        if self.base_df_all is None or self.rules is None:
            messagebox.showwarning("Base", "Primero carga base_comisiones y esquema.")
            return

        self.btn_process.configure(state="disabled")
        self.btn_export.configure(state="disabled")
        self._start_busy("Procesando comisiones...")

        threading.Thread(target=self._process_thread, daemon=True).start()

    def _process_thread(self):
        try:
            df = self.base_df_all.copy()

            # Fechas seleccionadas (si se usa filtro)
            d_ini = pd.to_datetime(self.cb_ini.get())
            d_fin = pd.to_datetime(self.cb_fin.get())
            if d_ini > d_fin:
                d_ini, d_fin = d_fin, d_ini

            # filtro opcional por fecha (si NO, solo etiqueta de carátula)
            if self.var_filter_date.get():
                self._set_status("Filtrando por fechas (opción activada)...")
                df = df[(df["Fecha"] >= d_ini) & (df["Fecha"] <= d_fin)].copy()
                if df.empty:
                    raise ValueError("No hay filas después de filtrar por fecha. Desactiva el filtro o cambia rango.")

            self._set_status("Calculando neto y venta total...")
            df["Asesor"] = df["Asesor"].fillna("").astype(str).str.strip()
            df["Cliente"] = df["Cliente"].fillna("").astype(str).str.strip()

            # Cálculo pedido: neto = bruto * 1.16
            df["Precio Unitario Neto"] = df["Precio Bruto"] * IVA_FACTOR
            df["Venta Total"] = df["Precio Unitario Neto"] * df["Cantidad"]

            self._set_status("Buscando precios 4–1 por producto...")
            prices_df = pd.DataFrame.from_dict(self.rules.price_map, orient="index")
            prices_df.index.name = "Producto_key"
            prices_df = prices_df.rename(columns={"p4": "Precio 4", "p3": "Precio 3", "p2": "Precio 2", "p1": "Precio 1"})

            df = df.merge(prices_df, left_on="Producto_key", right_index=True, how="left")

            self._set_status("Calculando bracket por asesor (ventas acumuladas)...")
            ventas_por_asesor = df.groupby("Asesor")["Venta Total"].sum().to_dict()
            bracket_by_asesor = {a: pick_bracket(self.rules.comm_brackets, float(v)) for a, v in ventas_por_asesor.items()}

            self._set_status("Calculando comisión por fila (por nivel de precio)...")
            compare_net = self.var_compare_net.get()

            # price to compare: neto o bruto
            if compare_net:
                price_for_level = df["Precio Unitario Neto"]
            else:
                price_for_level = df["Precio Bruto"]

            # vectorizar nivel
            def _infer_row(i):
                return infer_level(
                    float(price_for_level.iat[i]),
                    float(df["Precio 4"].iat[i]),
                    float(df["Precio 3"].iat[i]),
                    float(df["Precio 2"].iat[i]),
                    float(df["Precio 1"].iat[i]),
                )

            # Para performance: numpy loop simple
            levels = np.empty(len(df), dtype=int)
            p4 = df["Precio 4"].to_numpy()
            p3 = df["Precio 3"].to_numpy()
            p2 = df["Precio 2"].to_numpy()
            p1 = df["Precio 1"].to_numpy()
            pc = price_for_level.to_numpy()

            for i in range(len(df)):
                v = pc[i]
                if np.isnan(v) or np.isnan(p4[i]) or np.isnan(p3[i]) or np.isnan(p2[i]) or np.isnan(p1[i]):
                    levels[i] = 4
                else:
                    if v <= p4[i]:
                        levels[i] = 4
                    elif v <= p3[i]:
                        levels[i] = 3
                    elif v <= p2[i]:
                        levels[i] = 2
                    else:
                        levels[i] = 1

            df["Nivel"] = levels

            # rate por asesor según nivel
            asesor_arr = df["Asesor"].to_numpy()

            rates = np.empty(len(df), dtype=float)
            for i in range(len(df)):
                b = bracket_by_asesor.get(asesor_arr[i])
                if b is None:
                    rates[i] = 0.0
                else:
                    lvl = int(levels[i])
                    if lvl == 1:
                        rates[i] = b.p1
                    elif lvl == 2:
                        rates[i] = b.p2
                    elif lvl == 3:
                        rates[i] = b.p3
                    else:
                        rates[i] = b.p4

            df["Comisión"] = rates
            df["Total comisión"] = df["Comisión"] * df["Venta Total"]

            self._set_status("Preparando tablas y paginación...")
            out = df[DISPLAY_COLS].copy()

            resumen = out.groupby("Asesor", dropna=False).agg({
                "Venta Total": "sum",
                "Total comisión": "sum"
            }).reset_index().sort_values("Asesor")

            # Tipo opcional (si existe en Catalogos y se pidió para PDF)
            if self.var_pdf_tipo.get() and getattr(self.rules, "tipo_map", None):
                tipo_map = self.rules.tipo_map
                resumen["Tipo"] = resumen["Asesor"].apply(lambda a: tipo_map.get(str(a).strip().upper(), ""))
                # ordena para que Tipo quede después de Asesor
                cols = ["Asesor", "Tipo", "Venta Total", "Total comisión"]
                resumen = resumen[cols]

            def ui_apply():
                self.out_df = out
                self.resumen_df = resumen

                self._reset_pagination()
                self._render_page()
                self._render_resumen()
                self._render_audit_text(extra={
                    "rows_after_date_filter": int(len(df)),
                    "filter_by_date": bool(self.var_filter_date.get()),
                    "compare_price_by": "NETO" if compare_net else "BRUTO",
                })

                self.btn_export.configure(state="normal")
                self._stop_busy("Listo. Comisiones calculadas.")
                self.btn_process.configure(state="normal")

            self.after(0, ui_apply)

        except Exception as ex:
            def ui_err():
                self._stop_busy("Error al procesar.")
                self.btn_process.configure(state="normal")
                self.btn_export.configure(state="disabled")
                messagebox.showerror("Procesar", str(ex))
            self.after(0, ui_err)

    # ------------------ Auditoría ------------------
    def _render_audit_text(self, extra: Optional[Dict[str, object]] = None):
        self.txt_audit.delete("1.0", "end")
        lines = []
        lines.append("AUDITORÍA\n")
        if self.base_file:
            lines.append(f"Base: {self.base_file}")
        if self.schema_file:
            lines.append(f"Esquema: {self.schema_file}")
        lines.append("")
        for k, v in (self.audit_counts or {}).items():
            lines.append(f"- {k}: {v}")
        if extra:
            lines.append("")
            for k, v in extra.items():
                lines.append(f"- {k}: {v}")

        # Nota de consistencia
        lines.append("\nNOTAS:")
        lines.append("1) Detalle y Carátula SIEMPRE se calculan del MISMO dataset filtrado en esta app.")
        lines.append("2) Si tu Excel filtra por OVs (Hoja2) pero NO por Fecha, desactiva 'Filtrar ventas por Fecha'.")
        lines.append("3) Se excluyen líneas cuyo Producto contenga 'IVA', 'Impuesto' o 'Tax', y productos fuera de NUEVAS LISTAS.")

        self.txt_audit.insert("1.0", "\n".join(lines))

    # ------------------ Tablas ------------------
    def _clear_tables(self):
        for i in self.tree.get_children():
            self.tree.delete(i)
        for i in self.tree_sum.get_children():
            self.tree_sum.delete(i)

    def _render_resumen(self):
        for i in self.tree_sum.get_children():
            self.tree_sum.delete(i)

        if self.resumen_df is None:
            return

        for _, r in self.resumen_df.iterrows():
            self.tree_sum.insert("", "end", values=[
                r["Asesor"],
                money(r["Venta Total"]),
                money(r["Total comisión"])
            ])

    # ------------------ Paginación ------------------
    def _reset_pagination(self):
        self.current_page = 1
        n = 0 if self.out_df is None else len(self.out_df)
        self.total_pages = max(1, int(np.ceil(n / float(self.page_size))) if n else 1)
        self._update_pager_buttons()

    def _update_pager_buttons(self):
        n = 0 if self.out_df is None else len(self.out_df)
        has_data = n > 0
        self.lbl_page.configure(text=f"Página {self.current_page} / {self.total_pages}")

        state_enabled = "normal" if has_data and self.total_pages > 1 else "disabled"
        self.btn_first.configure(state=state_enabled)
        self.btn_prev.configure(state=state_enabled)
        self.btn_next.configure(state=state_enabled)
        self.btn_last.configure(state=state_enabled)

        if not has_data or self.total_pages <= 1:
            return
        if self.current_page <= 1:
            self.btn_first.configure(state="disabled")
            self.btn_prev.configure(state="disabled")
        if self.current_page >= self.total_pages:
            self.btn_next.configure(state="disabled")
            self.btn_last.configure(state="disabled")

    def _render_page(self):
        for i in self.tree.get_children():
            self.tree.delete(i)

        if self.out_df is None or self.out_df.empty:
            self._update_pager_buttons()
            return

        start = (self.current_page - 1) * self.page_size
        end = start + self.page_size
        view = self.out_df.iloc[start:end]

        for _, r in view.iterrows():
            vals = []
            for c in DISPLAY_COLS:
                v = r.get(c, "")
                if c in ("Precio Bruto", "Precio Unitario Neto", "Venta Total",
                         "Precio 4", "Precio 3", "Precio 2", "Precio 1", "Total comisión"):
                    vals.append(money(v))
                elif c == "Comisión":
                    vals.append(pct(v))
                elif c == "Fecha":
                    try:
                        vals.append(pd.to_datetime(v).strftime("%Y-%m-%d"))
                    except Exception:
                        vals.append("" if pd.isna(v) else str(v))
                else:
                    vals.append("" if pd.isna(v) else str(v))
            self.tree.insert("", "end", values=vals)

        self._update_pager_buttons()

    def on_pagesize_change(self, _evt=None):
        try:
            self.page_size = int(self.cb_pagesize.get())
        except Exception:
            self.page_size = DEFAULT_PAGE_SIZE
            self.cb_pagesize.set(str(DEFAULT_PAGE_SIZE))
        self._reset_pagination()
        self._render_page()

    def go_first(self):
        self.current_page = 1
        self._render_page()

    def go_last(self):
        self.current_page = self.total_pages
        self._render_page()

    def go_prev(self):
        if self.current_page > 1:
            self.current_page -= 1
            self._render_page()

    def go_next(self):
        if self.current_page < self.total_pages:
            self.current_page += 1
            self._render_page()

    # ------------------ Export PDF ------------------
    def export_pdf(self):
        if self.resumen_df is None:
            messagebox.showwarning("PDF", "Primero procesa para generar el resumen.")
            return

        d_ini = pd.to_datetime(self.cb_ini.get())
        d_fin = pd.to_datetime(self.cb_fin.get())
        d_pago = pd.to_datetime(self.cb_pago.get())

        p = filedialog.asksaveasfilename(
            title="Guardar carátula PDF",
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")]
        )
        if not p:
            return

        try:
            incluir_tipo = bool(self.var_pdf_tipo.get())
            resumen = self.resumen_df.copy()
            if incluir_tipo and "Tipo" not in resumen.columns and getattr(self.rules, "tipo_map", None):
                tipo_map = self.rules.tipo_map
                resumen["Tipo"] = resumen["Asesor"].apply(lambda a: tipo_map.get(str(a).strip().upper(), ""))

            export_caratula_pdf(
                p,
                resumen,
                d_ini,
                d_fin,
                d_pago,
                incluir_tipo=incluir_tipo
            )
            messagebox.showinfo("PDF", f"Carátula exportada:\n{p}")
        except Exception as ex:
            messagebox.showerror("PDF", str(ex))


if __name__ == "__main__":
    # Requisitos:
    #   pip install pandas numpy openpyxl reportlab
    app = Comisionador2026App()
    app.mainloop()
